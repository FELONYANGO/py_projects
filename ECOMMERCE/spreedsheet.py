import openpyxl as xl
from openpyxl import chart
from openpyxl.chart import BarChart, Reference


def process_Loader(filename):
    wk = xl.load_workbook(filename)

    sheet = wk["Sheet1"]
    cell= sheet["a1"]

    print(cell.value)

    for  row in range(2,sheet.max_row + 1):#looping through each row in the sheet 1
    
        cell = sheet.cell(row  ,3)
    
        corected_price = float(cell.value)*0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value=corected_price

    values=Reference(sheet,
            min_row = 2,
            max_row = sheet.max_row,
            min_col = 4,
            max_col=4)


    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"e2")


    wk.save(filename)